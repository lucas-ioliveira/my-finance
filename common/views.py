from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.views import View
from django.contrib import messages

from wallet.models import Expense, Revenue, Investments

import openpyxl

def custom_404(request, exception):
    return render(request, '404.html', status=404)

class ReportView(View):
    def get(self, request):
        headers = []
        filters = {
            'user': request.user,
            'active': True
        }

        model = request.GET.get('model')
        import ipdb; ipdb.set_trace()

        if model == 'Expense':

            if request.GET.get('status_exportar') != 'todos':
                filters['status'] = request.GET.get('status_exportar')
            
            if request.GET.get('categoria_exportar') != 'todos':
                filters['category'] = request.GET.get('categoria_exportar')
            
            if request.GET.get('data_vencimento_exportar'):
                filters['due_date__icontains'] = request.GET.get('data_vencimento_exportar')
            
            if request.GET.get('data_pagamento_exportar'):
                filters['payment_date__icontains'] = request.GET.get('data_pagamento_exportar')
            
            if request.GET.get('forma_pagamento_exportar'):
                filters['payment_method__icontains'] = request.GET.get('forma_pagamento_exportar')
            
            headers = ['título', 'observação', 'valor', 'categoria', 'data de pagamento', 'método de pagamento', 'status', 'data de vencimento']

            expense = Expense.objects.filter(**filters)

            if not expense.exists():
                messages.error(request, "Nenhuma despesa encontrada!")
                return redirect('expense')

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'relatorio_despesas'
            ws.append(headers)

            for i in expense:
                category_name = i.category.name if i.category else 'Sem Categoria'
                ws.append([i.description, i.notes, i.amount, category_name, i.payment_date, i.payment_method, i.status, i.due_date])

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=relatorio_despesas.xlsx'

            try:
                wb.save(response)
                return response
            except Exception as e:
                messages.error(request, f"Erro ao gerar o arquivo: {str(e)}")
                return redirect('expense')
        
        if model == 'Revenue':

            if request.GET.get('status_exportar') != 'todos':
                filters['status'] = request.GET.get('status_exportar')
            
            if request.GET.get('categoria_exportar') != 'todos':
                filters['category'] = request.GET.get('categoria_exportar')
    
            if request.GET.get('data_pagamento_exportar'):
                filters['payment_date__icontains'] = request.GET.get('data_pagamento_exportar')
            
            if request.GET.get('forma_pagamento_exportar'):
                filters['payment_method__icontains'] = request.GET.get('forma_pagamento_exportar')
            
            headers = ['título', 'observação', 'valor', 'categoria', 'data de pagamento', 'método de pagamento', 'status']

            revenue = Revenue.objects.filter(**filters)

            if not revenue.exists():
                messages.error(request, "Nenhuma despesa encontrada!")
                return redirect('revenue')

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'relatorio_receitas'
            ws.append(headers)

            for i in revenue:
                category_name = i.category.name if i.category else 'Sem Categoria'
                ws.append([i.description, i.notes, i.amount, category_name, i.payment_date, i.payment_method, i.status])

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=relatorio_receitas.xlsx'

            try:
                wb.save(response)
                return response
            except Exception as e:
                messages.error(request, f"Erro ao gerar o arquivo: {str(e)}")
                return redirect('revenue')
        
        if model == 'Investments':

            if request.GET.get('status_exportar') != 'todos':
                filters['status'] = request.GET.get('status_exportar')
            
            if request.GET.get('categoria_exportar') != 'todos':
                filters['category'] = request.GET.get('categoria_exportar')
    
            if request.GET.get('data_pagamento_exportar'):
                filters['investment_date__icontains'] = request.GET.get('data_pagamento_exportar')
            
            if request.GET.get('forma_pagamento_exportar'):
                filters['investment_method__icontains'] = request.GET.get('forma_pagamento_exportar')
            
            headers = ['título', 'observação', 'valor', 'categoria', 'data do investimento', 'método de investimento', 'status']

            investments = Investments.objects.filter(**filters)

            if not investments.exists():
                messages.error(request, "Nenhuma investimento encontrado!")
                return redirect('investments')

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'relatorio_investimentos'
            ws.append(headers)

            for i in investments:
                category_name = i.category.name if i.category else 'Sem Categoria'
                ws.append([i.description, i.notes, i.amount, category_name, i.investment_date, i.investment_method, i.status])

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=relatorio_investimentos.xlsx'

            try:
                wb.save(response)
                return response
            except Exception as e:
                messages.error(request, f"Erro ao gerar o arquivo: {str(e)}")
                return redirect('investments')
            
        else:
            messages.error(request, "Erro ao gerar o arquivo!")
            return redirect('investments')
        