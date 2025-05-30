import openpyxl
from django.http import HttpResponse

from wallet.repositories.expense.repository import ExpenseRepository
from wallet.repositories.investments.repository import InvestmentsRepository
from wallet.repositories.revanue.repository import RevenueRepository


class CommonService:

    @staticmethod
    def generate_report(request):
        headers = []
        filters = {
            'user': request.user.id,
            'active': True
        }

        model = request.GET.get('model')

        if model == 'Expense':

            if request.GET.get('status_exportar') != 'todos':
                filters['status'] = request.GET.get('status_exportar')

            if request.GET.get('categoria_exportar') != 'todos':
                filters['category'] = request.GET.get('categoria_exportar')

            if request.GET.get('data_vencimento_exportar_inicial'):
                filters['due_date__gte'] = request.GET.get('data_vencimento_exportar_inicial')

            if request.GET.get('data_vencimento_exportar_final'):
                filters['due_date__lte'] = request.GET.get('data_vencimento_exportar_final')

            if request.GET.get('data_pagamento_exportar_inicial'):
                filters['payment_date__gte'] = request.GET.get('data_pagamento_exportar_inicial')

            if request.GET.get('data_pagamento_exportar_final'):
                filters['payment_date__lte'] = request.GET.get('data_pagamento_exportar_final')

            if request.GET.get('forma_pagamento_exportar'):
                filters['payment_method__icontains'] = request.GET.get('forma_pagamento_exportar')

            headers = ['título', 'observação', 'valor', 'categoria', 'data de pagamento', 'método de pagamento', 'status', 'data de vencimento']

            expense = ExpenseRepository.get_by_filters_for_the_report(**filters)

            if not expense.exists():
                raise Exception('Nenhuma despesa encontrada!')

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'relatorio_despesas'
            ws.append(headers)

            for i in expense:
                category_name = i.category.name if i.category else 'Sem Categoria'
                ws.append([i.description, i.notes, i.amount, category_name, i.payment_date, i.payment_method, i.status, i.due_date])

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=relatorio_despesas.xlsx'

            try:
                wb.save(response)
                return response
            except Exception as e:

                raise Exception(f"Erro ao gerar o arquivo: {str(e)}")

        if model == 'Revenue':

            if request.GET.get('status_exportar') != 'todos':
                filters['status'] = request.GET.get('status_exportar')

            if request.GET.get('categoria_exportar') != 'todos':
                filters['category'] = request.GET.get('categoria_exportar')

            if request.GET.get('data_pagamento_exportar_inicial'):
                filters['payment_date__gte'] = request.GET.get('data_pagamento_exportar_inicial')

            if request.GET.get('data_pagamento_exportar_final'):
                filters['payment_date__lte'] = request.GET.get('data_pagamento_exportar_final')

            if request.GET.get('forma_pagamento_exportar'):
                filters['payment_method__icontains'] = request.GET.get('forma_pagamento_exportar')

            headers = ['título', 'observação', 'valor', 'categoria', 'data de pagamento', 'método de pagamento', 'status']

            revenue = RevenueRepository.get_by_filters_for_the_report(**filters)

            if not revenue.exists():

                raise Exception('Nenhuma receita encontrada!')

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'relatorio_receitas'
            ws.append(headers)

            for i in revenue:
                category_name = i.category.name if i.category else 'Sem Categoria'
                ws.append([i.description, i.notes, i.amount, category_name, i.payment_date, i.payment_method, i.status])

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=relatorio_receitas.xlsx'

            try:
                wb.save(response)
                return response
            except Exception as e:

                raise Exception(f"Erro ao gerar o arquivo: {str(e)}")

        if model == 'Investments':

            if request.GET.get('status_exportar') != 'todos':
                filters['status'] = request.GET.get('status_exportar')

            if request.GET.get('categoria_exportar') != 'todos':
                filters['category'] = request.GET.get('categoria_exportar')

            if request.GET.get('data_investimento_exportar_inicial'):
                filters['investment_date__gte'] = request.GET.get('data_investimento_exportar_inicial')

            if request.GET.get('data_investimento_exportar_final'):
                filters['investment_date__lte'] = request.GET.get('data_investimento_exportar_final')

            if request.GET.get('forma_pagamento_exportar'):
                filters['investment_method__icontains'] = request.GET.get('forma_pagamento_exportar')

            headers = ['título', 'observação', 'valor', 'categoria', 'data do investimento', 'método de investimento', 'status']

            investments = InvestmentsRepository.get_by_filters_for_the_report(**filters)

            if not investments.exists():

                raise Exception('Nenhum investimento encontrado!')

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'relatorio_investimentos'
            ws.append(headers)

            for i in investments:
                category_name = i.category.name if i.category else 'Sem Categoria'
                ws.append([i.description, i.notes, i.amount, category_name, i.investment_date, i.investment_method, i.status])

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=relatorio_investimentos.xlsx'

            try:
                wb.save(response)
                return response
            except Exception as e:

                raise Exception(f"Erro ao gerar o arquivo: {str(e)}")

        return Exception('Erro ao gerar o arquivo!')
